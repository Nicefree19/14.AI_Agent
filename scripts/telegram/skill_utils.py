#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
스킬 공통 유틸리티 모듈

볼트 이슈 로딩, PDF 처리, SEN 패턴 감지, COM 헬퍼 등
모든 스킬 모듈에서 공유하는 함수 모음.
"""

from __future__ import annotations

import os
import re
import time
import traceback
from pathlib import Path
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

import yaml

# ─── 경로 설정 ───────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent          # scripts/telegram/
SCRIPTS_DIR = SCRIPT_DIR.parent                        # scripts/
PROJECT_ROOT = SCRIPTS_DIR.parent                      # 14.AI_Agent/
VAULT_PATH = PROJECT_ROOT / "ResearchVault"
ISSUES_DIR = VAULT_PATH / "P5-Project" / "01-Issues"
CONFIG_DIR = VAULT_PATH / "_config"
OVERVIEW_DIR = VAULT_PATH / "P5-Project" / "00-Overview"

# ─── P5 도면번호 정규식 (drawing_extractor.py 패턴 재사용) ────
DRAWING_PATTERNS = [
    re.compile(r"\b([A-Z]{2,4}-\d{3,})\b"),
    re.compile(r"\b(S-\d{3,})\b"),
    re.compile(r"\b(SHOP[-_]R(?:ev)?\d+)\b", re.IGNORECASE),
    re.compile(r"\b(EP[-_]\d{2,})\b"),
    re.compile(r"\b(PSRC[-_]\d{2,})\b"),
    re.compile(r"\b(HMB[-_]\d{2,})\b"),
    re.compile(r"\b(PLE[GB][-_]\d{2,})\b"),
    re.compile(r"\b(FCC[-_]\d{2,})\b"),
    re.compile(r"\b(D[WR][GW][-_]\d{3,})\b", re.IGNORECASE),
]

# SEN 이슈 ID 패턴
SEN_PATTERN = re.compile(r"\b(SEN[-_]\d{3,})\b", re.IGNORECASE)

# ─── 카테고리/우선순위 매핑 (p5_risk_matrix.py 참조) ─────────
PRIORITY_URGENCY = {
    "critical": 0.9,
    "high": 0.7,
    "medium": 0.5,
    "normal": 0.4,
    "low": 0.2,
}

CATEGORY_IMPACT = {
    "구조접합": 0.9,
    "간섭": 0.85,
    "psrc": 0.85,
    "hmb": 0.8,
    "설계": 0.75,
    "pc연동": 0.7,
    "일정": 0.6,
    "상세변경": 0.5,
    "해당없음": 0.3,
}

# 제작 관련 카테고리
FABRICATION_CATEGORIES = {"psrc", "hmb", "구조접합", "pc연동", "상세변경"}

# 제작 단계 키워드 매핑
FABRICATION_STAGES = {
    "설계검토": ["설계", "검토", "계산서", "확인"],
    "Shop DWG": ["shop", "드로잉", "도면", "dwg"],
    "제작중": ["제작", "가공", "용접", "볼트"],
    "납품": ["납품", "운송", "출하", "반입"],
    "시공": ["시공", "설치", "양중", "조립"],
}


def classify_stages(issues: List[Dict]) -> Dict[str, List[Dict]]:
    """이슈를 제작 단계별로 분류. (공유 헬퍼)"""
    stages: Dict[str, List[Dict]] = {stage: [] for stage in FABRICATION_STAGES}
    stages["미분류"] = []

    for issue in issues:
        text = f"{issue.get('title', '')} {issue.get('_body', '')} {issue.get('status', '')}".lower()
        classified = False
        for stage_name, keywords in FABRICATION_STAGES.items():
            if any(kw in text for kw in keywords):
                stages[stage_name].append(issue)
                classified = True
                break
        if not classified:
            stages["미분류"].append(issue)

    return {k: v for k, v in stages.items() if v}


def get_stage_icon(stage: str) -> str:
    """단계별 아이콘. (공유 헬퍼)"""
    icons = {
        "설계검토": "📋",
        "Shop DWG": "📐",
        "제작중": "🏭",
        "납품": "🚚",
        "시공": "🏗️",
        "미분류": "❓",
    }
    return icons.get(stage, "•")


# ═══════════════════════════════════════════════════════════════
#  1. 볼트 이슈 로딩 (TTL 캐시 적용)
# ═══════════════════════════════════════════════════════════════

# 모듈 레벨 캐시 — 동일 프로세스 내 중복 디스크 I/O 방지
_vault_issues_cache: List[Dict] = []
_vault_issues_cache_time: float = 0.0
_VAULT_CACHE_TTL: float = 300.0  # 5분


def _load_vault_issues_raw() -> List[Dict]:
    """디스크에서 모든 이슈를 읽어온다 (캐시 미적용)."""
    if not ISSUES_DIR.exists():
        return []

    issues: List[Dict] = []
    for f in ISSUES_DIR.glob("*.md"):
        if f.name.startswith("20"):  # 인덱스 파일 제외
            continue
        try:
            content = f.read_text(encoding="utf-8")
            if not content.startswith("---"):
                continue
            parts = content.split("---", 2)
            if len(parts) >= 2:
                fm = yaml.safe_load(parts[1]) or {}
                if fm.get("issue_id"):
                    fm["_file_path"] = str(f)
                    fm["_body"] = parts[2].strip() if len(parts) > 2 else ""
                    issues.append(fm)
        except Exception:
            pass

    return issues


def load_vault_issues(
    filters: Optional[Dict[str, Any]] = None,
) -> List[Dict]:
    """
    볼트 이슈 YAML frontmatter 로딩 (TTL 캐시 적용).

    동일 프로세스 내에서 5분 이내 재호출 시 캐시된 결과를 반환한다.
    여러 스킬이 한 번의 실행에서 이슈를 반복 로딩하는 경우 디스크 I/O를 절약한다.

    Args:
        filters: 선택적 필터 dict. 지원 키:
            - category: str — 카테고리 필터 (부분 매칭)
            - priority: str|list — 우선순위 필터
            - status: str — 상태 필터 (open/closed/in_progress)
            - owner: str — 담당자 필터 (부분 매칭)
            - since_days: int — 최근 N일 이내 생성

    Returns:
        이슈 frontmatter dict 리스트
    """
    global _vault_issues_cache, _vault_issues_cache_time

    now = time.time()
    if not _vault_issues_cache or (now - _vault_issues_cache_time) > _VAULT_CACHE_TTL:
        _vault_issues_cache = _load_vault_issues_raw()
        _vault_issues_cache_time = now

    # 필터 없으면 캐시 전체 반환 (복사본)
    if not filters:
        return list(_vault_issues_cache)

    return _apply_issue_filters(list(_vault_issues_cache), filters)


def _apply_issue_filters(
    issues: List[Dict], filters: Dict[str, Any]
) -> List[Dict]:
    """이슈 필터 적용."""
    result = issues

    if "category" in filters:
        cat = filters["category"].lower()
        result = [i for i in result if cat in i.get("category", "").lower()]

    if "priority" in filters:
        prio = filters["priority"]
        if isinstance(prio, str):
            prio = [prio]
        prio_lower = [p.lower() for p in prio]
        result = [i for i in result if i.get("priority", "").lower() in prio_lower]

    if "status" in filters:
        status = filters["status"].lower()
        result = [i for i in result if i.get("status", "").lower() == status]

    if "owner" in filters:
        owner = filters["owner"]
        result = [
            i for i in result
            if owner in i.get("owner", "") or owner in i.get("assignee", "")
        ]

    if "since_days" in filters:
        cutoff = datetime.now() - timedelta(days=filters["since_days"])
        cutoff_str = cutoff.strftime("%Y-%m-%d")
        result = [
            i for i in result
            if i.get("created", "") >= cutoff_str or i.get("date", "") >= cutoff_str
        ]

    return result


def search_issues(keyword: str, max_results: int = 10) -> List[Dict]:
    """
    이슈 전문 키워드 검색 (제목, 설명, 카테고리, 조치계획).

    Args:
        keyword: 검색 키워드
        max_results: 최대 결과 수

    Returns:
        매칭된 이슈 리스트 (relevance_score 포함)
    """
    all_issues = load_vault_issues()
    keyword_lower = keyword.lower()
    scored: List[Tuple[float, Dict]] = []

    for issue in all_issues:
        score = 0.0
        # issue_id 정확 매칭
        if keyword_lower in issue.get("issue_id", "").lower():
            score += 10.0
        # 제목 매칭
        if keyword_lower in issue.get("title", "").lower():
            score += 5.0
        # 카테고리 매칭
        if keyword_lower in issue.get("category", "").lower():
            score += 3.0
        # 담당자 매칭
        if keyword_lower in issue.get("owner", "").lower():
            score += 2.0
        # 본문 매칭
        if keyword_lower in issue.get("_body", "").lower():
            score += 1.0
        # 조치계획 매칭
        if keyword_lower in str(issue.get("action_plan", "")).lower():
            score += 2.0

        if score > 0:
            issue["_relevance"] = score
            scored.append((score, issue))

    scored.sort(key=lambda x: x[0], reverse=True)
    return [item[1] for item in scored[:max_results]]


def get_issue_by_id(issue_id: str) -> Optional[Dict]:
    """정확한 이슈 ID로 조회."""
    issue_id_upper = issue_id.upper().replace("_", "-")
    all_issues = load_vault_issues()
    for issue in all_issues:
        if issue.get("issue_id", "").upper() == issue_id_upper:
            return issue
    return None


def format_issue_detail(issue: Dict) -> str:
    """이슈를 읽기 좋은 텍스트로 포맷팅."""
    lines = []
    iid = issue.get("issue_id", "N/A")
    title = issue.get("title", "제목 없음")
    lines.append(f"📌 {iid}: {title}")
    lines.append(f"━━━━━━━━━━━━━━━━━━━━━━━━")

    fields = [
        ("카테고리", "category"),
        ("우선순위", "priority"),
        ("상태", "status"),
        ("담당자", "owner"),
        ("마감일", "due_date"),
        ("생성일", "created"),
    ]
    for label, key in fields:
        val = issue.get(key)
        if val:
            lines.append(f"• {label}: {val}")

    action = issue.get("action_plan")
    if action:
        lines.append(f"\n📋 조치계획:")
        if isinstance(action, list):
            for item in action:
                lines.append(f"  → {item}")
        else:
            lines.append(f"  → {action}")

    body = issue.get("_body", "").strip()
    if body:
        # 최대 500자
        preview = body[:500] + ("..." if len(body) > 500 else "")
        lines.append(f"\n📝 내용:\n{preview}")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════
#  2. 파일 추출 유틸리티
# ═══════════════════════════════════════════════════════════════

def extract_files_by_ext(
    context: dict, extensions: List[str]
) -> List[Dict]:
    """
    context의 첨부 파일에서 특정 확장자 파일 추출.

    Args:
        context: executor context dict
        extensions: 추출할 확장자 리스트 (e.g., [".pdf", ".PDF"])

    Returns:
        매칭된 파일 정보 dict 리스트
    """
    combined = context.get("combined", {})
    files = combined.get("files", [])
    ext_lower = {e.lower() for e in extensions}

    matched = []
    for f in files:
        name = f.get("name", "")
        _, ext = os.path.splitext(name)
        if ext.lower() in ext_lower:
            matched.append(f)
    return matched


def get_file_path(file_info: dict, task_dir: str) -> Optional[str]:
    """파일 정보에서 실제 경로 결정."""
    # local_path가 있으면 사용
    local = file_info.get("local_path")
    if local and os.path.exists(local):
        return local
    # task_dir에서 파일명으로 검색
    name = file_info.get("name", "")
    if name:
        candidate = os.path.join(task_dir, name)
        if os.path.exists(candidate):
            return candidate
    return None


# ═══════════════════════════════════════════════════════════════
#  3. PDF 유틸리티
# ═══════════════════════════════════════════════════════════════

def extract_pdf_text(
    path: str, max_pages: int = 50
) -> Tuple[str, int]:
    """
    PDF에서 텍스트 추출.

    Args:
        path: PDF 파일 경로
        max_pages: 최대 처리 페이지 수

    Returns:
        (추출된 텍스트, 총 페이지수)
    """
    try:
        import fitz  # pymupdf
    except ImportError:
        return "⚠️ pymupdf가 설치되지 않았습니다.", 0

    text_parts: List[str] = []
    try:
        doc = fitz.open(path)
        total_pages = len(doc)
        pages_to_read = min(total_pages, max_pages)

        for i in range(pages_to_read):
            page = doc[i]
            page_text = page.get_text()
            if page_text.strip():
                text_parts.append(f"--- 페이지 {i + 1}/{total_pages} ---")
                text_parts.append(page_text.strip())

        doc.close()

        if total_pages > max_pages:
            text_parts.append(
                f"\n⚠️ 총 {total_pages}페이지 중 {max_pages}페이지까지만 추출했습니다."
            )

        return "\n\n".join(text_parts), total_pages

    except Exception as e:
        return f"❌ PDF 텍스트 추출 오류: {e}", 0


def render_pdf_page(
    path: str, page: int = 0, dpi: int = 150
) -> Optional[str]:
    """
    PDF 페이지를 PNG 이미지로 렌더링.

    Args:
        path: PDF 파일 경로
        page: 페이지 번호 (0-based)
        dpi: 해상도

    Returns:
        저장된 이미지 경로 또는 None
    """
    try:
        import fitz
    except ImportError:
        return None

    try:
        doc = fitz.open(path)
        if page >= len(doc):
            doc.close()
            return None

        pg = doc[page]
        mat = fitz.Matrix(dpi / 72, dpi / 72)
        pix = pg.get_pixmap(matrix=mat)

        output_path = Path(path).with_suffix(f".page{page + 1}.png")
        pix.save(str(output_path))
        doc.close()
        return str(output_path)

    except Exception:
        return None


def detect_pdf_structure(text: str) -> Dict[str, Any]:
    """
    PDF 텍스트에서 구조 감지 (헤더, 표, 목록).

    Returns:
        {"headers": list, "tables_hint": bool, "list_items": int, "drawing_refs": list}
    """
    lines = text.split("\n")
    headers = []
    list_count = 0
    table_hint = False

    for line in lines:
        stripped = line.strip()
        # 헤더 감지 (짧은 줄 + 대문자 비율 높음)
        if 3 < len(stripped) < 80:
            upper_ratio = sum(1 for c in stripped if c.isupper()) / max(len(stripped), 1)
            if upper_ratio > 0.5:
                headers.append(stripped)
        # 목록 감지
        if stripped.startswith(("•", "-", "·", "※", "▪")) or re.match(r"^\d+[\.\)]\s", stripped):
            list_count += 1
        # 표 감지 (탭 또는 여러 공백 구분자)
        if "\t" in stripped or "  |  " in stripped or stripped.count("|") > 2:
            table_hint = True

    # 도면번호 추출
    drawing_refs = detect_drawing_refs(text)

    return {
        "headers": headers[:20],
        "tables_hint": table_hint,
        "list_items": list_count,
        "drawing_refs": drawing_refs,
    }


# ═══════════════════════════════════════════════════════════════
#  4. 패턴 감지
# ═══════════════════════════════════════════════════════════════

def detect_sen_refs(text: str) -> List[str]:
    """텍스트에서 SEN-XXX 이슈 ID 추출."""
    matches = SEN_PATTERN.findall(text)
    # 정규화: SEN_001 → SEN-001
    normalized = sorted(set(m.upper().replace("_", "-") for m in matches))
    return normalized


def detect_drawing_refs(text: str) -> List[str]:
    """텍스트에서 도면번호 추출."""
    found: set = set()
    for pattern in DRAWING_PATTERNS:
        for match in pattern.finditer(text):
            found.add(match.group(1).upper())
    return sorted(found)


# ═══════════════════════════════════════════════════════════════
#  5. COM 헬퍼 (Excel / PPT)
# ═══════════════════════════════════════════════════════════════

def create_excel_workbook(
    sheets_data: Dict[str, Tuple[List[str], List[List[Any]]]],
    output_path: str,
    use_com: bool = True,
) -> Optional[str]:
    """
    Excel 워크북 생성.

    Args:
        sheets_data: {시트명: (헤더리스트, 데이터행리스트)}
        output_path: 출력 파일 경로
        use_com: True=COM 사용 (서식 지원), False=openpyxl만 사용

    Returns:
        생성된 파일 경로 또는 None
    """
    if use_com:
        result = _create_excel_com(sheets_data, output_path)
        if result:
            return result
        # COM 실패 시 openpyxl fallback

    return _create_excel_openpyxl(sheets_data, output_path)


def _create_excel_com(
    sheets_data: Dict[str, Tuple[List[str], List[List[Any]]]],
    output_path: str,
) -> Optional[str]:
    """COM으로 Excel 생성 (서식 포함)."""
    app = None
    try:
        import win32com.client
        import pythoncom

        pythoncom.CoInitialize()
        app = win32com.client.Dispatch("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False

        wb = app.Workbooks.Add()

        for idx, (sheet_name, (headers, rows)) in enumerate(sheets_data.items()):
            if idx == 0:
                ws = wb.Sheets(1)
                ws.Name = sheet_name[:31]  # Excel 시트명 31자 제한
            else:
                ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                ws.Name = sheet_name[:31]

            # 헤더
            for col, h in enumerate(headers, 1):
                cell = ws.Cells(1, col)
                cell.Value = h
                cell.Font.Bold = True
                cell.Interior.Color = 0xD9E2F3  # 연한 파란색

            # 데이터
            for row_idx, row in enumerate(rows, 2):
                for col_idx, val in enumerate(row, 1):
                    ws.Cells(row_idx, col_idx).Value = val

            # 열 너비 자동 조정
            ws.Columns.AutoFit()

        # 저장
        abs_path = os.path.abspath(output_path)
        wb.SaveAs(abs_path, FileFormat=51)  # 51 = xlsx
        wb.Close(SaveChanges=False)
        return abs_path

    except Exception:
        return None
    finally:
        if app:
            try:
                app.Quit()
            except Exception:
                pass
            try:
                import pythoncom
                pythoncom.CoUninitialize()
            except Exception:
                pass


def _create_excel_openpyxl(
    sheets_data: Dict[str, Tuple[List[str], List[List[Any]]]],
    output_path: str,
) -> Optional[str]:
    """openpyxl로 Excel 생성 (COM 불가 시 fallback)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()

        for idx, (sheet_name, (headers, rows)) in enumerate(sheets_data.items()):
            if idx == 0:
                ws = wb.active
                ws.title = sheet_name[:31]
            else:
                ws = wb.create_sheet(title=sheet_name[:31])

            # 헤더
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            # 데이터
            for row_idx, row in enumerate(rows, 2):
                for col_idx, val in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)

        wb.save(output_path)
        return os.path.abspath(output_path)

    except Exception:
        return None


def create_ppt_com(
    slides_data: List[Dict[str, Any]],
    output_path: str,
) -> Optional[str]:
    """
    COM으로 PowerPoint 생성.

    Args:
        slides_data: [{"title": str, "content": str, "layout": int}]
        output_path: 출력 파일 경로

    Returns:
        생성된 파일 경로 또는 None
    """
    app = None
    try:
        import win32com.client
        import pythoncom

        pythoncom.CoInitialize()
        app = win32com.client.Dispatch("PowerPoint.Application")

        prs = app.Presentations.Add(WithWindow=False)

        for slide_data in slides_data:
            layout_idx = slide_data.get("layout", 2)  # 2 = Title and Content
            layout = prs.SlideMaster.CustomLayouts(layout_idx)
            slide = prs.Slides.AddSlide(prs.Slides.Count + 1, layout)

            # 제목
            if slide.Shapes.HasTitle and slide_data.get("title"):
                slide.Shapes.Title.TextFrame.TextRange.Text = slide_data["title"]

            # 내용
            content = slide_data.get("content", "")
            if content and slide.Shapes.Count > 1:
                body = slide.Shapes(2)
                if body.HasTextFrame:
                    body.TextFrame.TextRange.Text = content

        abs_path = os.path.abspath(output_path)
        prs.SaveAs(abs_path)
        prs.Close()
        return abs_path

    except Exception:
        return None
    finally:
        if app:
            try:
                app.Quit()
            except Exception:
                pass
            try:
                import pythoncom
                pythoncom.CoUninitialize()
            except Exception:
                pass


def safe_com_cleanup(app: Any) -> None:
    """COM 객체 안전 정리."""
    if app is None:
        return
    try:
        app.Quit()
    except Exception:
        pass
    try:
        del app
    except Exception:
        pass
    try:
        import pythoncom
        pythoncom.CoUninitialize()
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════
#  6. 텍스트 유틸리티
# ═══════════════════════════════════════════════════════════════

def truncate_text(text: str, max_chars: int = 4000) -> str:
    """텍스트를 최대 길이로 잘라냄 (초과 시 파일 저장 안내)."""
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + f"\n\n... (총 {len(text):,}자 중 {max_chars:,}자만 표시)"


def save_text_to_file(text: str, task_dir: str, filename: str) -> str:
    """텍스트를 파일로 저장하고 경로 반환."""
    os.makedirs(task_dir, exist_ok=True)
    filepath = os.path.join(task_dir, filename)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(text)
    return filepath


def format_matrix_text(
    headers: List[str],
    rows: List[List[str]],
    title: str = "",
) -> str:
    """간단한 텍스트 매트릭스 포맷팅."""
    lines = []
    if title:
        lines.append(f"📊 {title}")
        lines.append("━" * 40)

    # 열 너비 계산
    widths = [len(h) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            if i < len(widths):
                widths[i] = max(widths[i], len(str(cell)))

    # 헤더
    header_line = " | ".join(h.ljust(widths[i]) for i, h in enumerate(headers))
    lines.append(header_line)
    lines.append("-" * len(header_line))

    # 데이터
    for row in rows:
        row_line = " | ".join(
            str(cell).ljust(widths[i]) if i < len(widths) else str(cell)
            for i, cell in enumerate(row)
        )
        lines.append(row_line)

    return "\n".join(lines)


def extract_instruction_keyword(instruction: str, keywords: List[str]) -> Optional[str]:
    """지시 텍스트에서 특정 키워드 추출."""
    inst_lower = instruction.lower()
    for kw in keywords:
        if kw.lower() in inst_lower:
            return kw
    return None


# ═══════════════════════════════════════════════════════════════
#  7. 이슈 매칭 유틸리티 (회의록·통화 등 범용)
# ═══════════════════════════════════════════════════════════════

def match_issues_by_topic(
    topic_text: str,
    all_issues: Optional[List[Dict]] = None,
    threshold: float = 0.65,
    max_results: int = 3,
) -> List[Dict]:
    """3단계 이슈 매칭 (exact ID → title similarity → keyword overlap).

    p5_email_triage.py TriageEngine._match_issue() 패턴 재사용.

    Args:
        topic_text: 매칭 대상 텍스트 (결정사항, 액션아이템 등)
        all_issues: 이슈 목록 (None이면 load_vault_issues() 호출)
        threshold: Tier 2 유사도 임계값 (회의는 0.65, 이메일은 0.7)
        max_results: 최대 결과 수

    Returns:
        [{"issue": dict, "confidence": float, "tier": int, "reason": str}, ...]
    """
    from difflib import SequenceMatcher

    if all_issues is None:
        all_issues = load_vault_issues()

    results: List[Dict] = []

    # ── Tier 1: SEN-xxx ID 직접 매칭 (confidence 1.0) ──
    sen_refs = detect_sen_refs(topic_text)
    for ref in sen_refs:
        issue = get_issue_by_id(ref)
        if issue:
            results.append({
                "issue": issue,
                "confidence": 1.0,
                "tier": 1,
                "reason": f"ID 직접 매칭: {ref}",
            })

    matched_ids = {r["issue"].get("issue_id") for r in results}

    # ── Tier 2: 제목 유사도 (SequenceMatcher) ──
    topic_lower = topic_text.lower()
    tier2_candidates: List[Tuple[float, Dict]] = []

    for issue in all_issues:
        iid = issue.get("issue_id", "")
        if iid in matched_ids:
            continue

        title = issue.get("title", "").lower()
        desc = issue.get("description", "").lower()
        action = str(issue.get("action_plan", "")).lower()

        # 제목·설명·조치계획 모두 비교, 최대값 사용
        scores = [
            SequenceMatcher(None, topic_lower, title).ratio(),
            SequenceMatcher(None, topic_lower, desc).ratio() * 0.8,
            SequenceMatcher(None, topic_lower, action).ratio() * 0.7,
        ]
        best_score = max(scores)
        if best_score >= threshold:
            tier2_candidates.append((best_score, issue))

    tier2_candidates.sort(key=lambda x: x[0], reverse=True)
    for score, issue in tier2_candidates[:max_results]:
        iid = issue.get("issue_id", "")
        if iid not in matched_ids:
            results.append({
                "issue": issue,
                "confidence": round(score, 2),
                "tier": 2,
                "reason": f"제목 유사도: {score:.0%}",
            })
            matched_ids.add(iid)

    # ── Tier 3: 카테고리 + 담당자 키워드 겹침 ──
    if len(results) < max_results:
        topic_words = set(re.findall(r"[\w가-힣]+", topic_lower))
        tier3_candidates: List[Tuple[float, Dict]] = []

        for issue in all_issues:
            iid = issue.get("issue_id", "")
            if iid in matched_ids:
                continue

            issue_words = set()
            for field in ("title", "category", "owner", "description"):
                val = str(issue.get(field, "")).lower()
                issue_words.update(re.findall(r"[\w가-힣]+", val))

            if not issue_words or not topic_words:
                continue

            overlap = len(topic_words & issue_words)
            jaccard = overlap / len(topic_words | issue_words) if topic_words | issue_words else 0
            if jaccard >= 0.15 and overlap >= 2:
                tier3_candidates.append((round(jaccard, 2), issue))

        tier3_candidates.sort(key=lambda x: x[0], reverse=True)
        for score, issue in tier3_candidates[:max_results - len(results)]:
            iid = issue.get("issue_id", "")
            if iid not in matched_ids:
                results.append({
                    "issue": issue,
                    "confidence": round(min(score * 1.5, 0.6), 2),
                    "tier": 3,
                    "reason": f"키워드 겹침: {score:.0%}",
                })
                matched_ids.add(iid)

    return results[:max_results]


def update_issue_field_append(
    filepath: str,
    field: str,
    new_value: str,
    separator: str = " | ",
) -> bool:
    """이슈 frontmatter 필드에 값 추가 (기존 값 보존).

    engineering_skills.py _update_issue_frontmatter() 패턴 재사용.

    Args:
        filepath: 이슈 .md 파일 경로
        field: frontmatter 필드명 (예: "related_docs", "decision")
        new_value: 추가할 값
        separator: 기존 값과 새 값 사이 구분자

    Returns:
        True: 업데이트 성공, False: 실패
    """
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()

        if not content.startswith("---"):
            return False

        parts = content.split("---", 2)
        if len(parts) < 3:
            return False

        fm_text = parts[1]
        pattern = rf"^({re.escape(field)}\s*:\s*)(.*)$"
        match = re.search(pattern, fm_text, re.MULTILINE)

        if match:
            existing = match.group(2).strip().strip('"').strip("'")
            if new_value in existing:
                return True  # 이미 존재
            if existing and existing != '""' and existing != "''":
                updated = f'{existing}{separator}{new_value}'
            else:
                updated = new_value
            new_line = f'{field}: "{updated}"'
            fm_text = re.sub(pattern, new_line, fm_text, flags=re.MULTILINE)
        else:
            fm_text = fm_text.rstrip() + f'\n{field}: "{new_value}"\n'

        new_content = f"---{fm_text}---{parts[2]}"
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(new_content)
        return True

    except Exception:
        return False
