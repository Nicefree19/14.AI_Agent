#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Analyze the 센코어테크 Excel attachment."""
import sys, os, io

if sys.stdout.encoding and sys.stdout.encoding.lower().startswith("cp"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(_ROOT)

EXCEL_PATH = os.path.join(_ROOT, "telegram_data", "tasks", "msg_38", "26.2.09 P5복합동 생산(출하)일보 - 센코어테크.xlsx")

import openpyxl

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
print(f"Workbook sheets: {wb.sheetnames}")
print(f"Total sheets: {len(wb.sheetnames)}")
print("=" * 100)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*100}")
    print(f"Sheet: '{sheet_name}'")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

    if ws.max_row == 0 or ws.max_column == 0:
        print("  (empty sheet)")
        continue

    # Read all non-empty rows (up to 200 rows for analysis)
    print(f"\n--- Data (up to 200 rows) ---")
    row_count = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), values_only=False):
        row_count += 1
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                vals.append(f"[{cell.coordinate}]{v}")
        if vals:
            # Print first 10 cells per row to avoid excessive output
            line = " | ".join(vals[:12])
            if len(vals) > 12:
                line += f" ... (+{len(vals)-12} more)"
            print(f"  Row {row_count}: {line[:300]}")

    # Check for merged cells
    if ws.merged_cells.ranges:
        print(f"\n  Merged cells: {len(ws.merged_cells.ranges)}")
        for mc in list(ws.merged_cells.ranges)[:20]:
            print(f"    {mc}")

    print()

wb.close()
