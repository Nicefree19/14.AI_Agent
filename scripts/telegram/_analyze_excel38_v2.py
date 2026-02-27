#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Focused analysis of the 센코어테크 Excel file - extract key production data."""
import sys, os, io

if sys.stdout.encoding and sys.stdout.encoding.lower().startswith("cp"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(_ROOT)

EXCEL_PATH = os.path.join(_ROOT, "telegram_data", "tasks", "msg_38", "26.2.09 P5복합동 생산(출하)일보 - 센코어테크.xlsx")

import openpyxl

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
print(f"=== WORKBOOK OVERVIEW ===")
print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: '{sheet_name}' | Rows: {ws.max_row} | Cols: {ws.max_column}")
    print(f"{'='*80}")

    if ws.max_row == 0 or ws.max_column == 0:
        print("  (empty)")
        continue

    # Print first 30 rows only (headers + key data)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=False), 1):
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                s = str(v)
                if len(s) > 50:
                    s = s[:50] + "..."
                vals.append(f"{cell.column_letter}{cell.row}={s}")
        if vals:
            print(f"  R{row_idx:3d}: {' | '.join(vals[:15])}")

    # If more than 30 rows, show a sample from middle and end
    if ws.max_row > 30:
        print(f"\n  ... ({ws.max_row - 30} more rows) ...")

        # Show last 5 rows
        print(f"\n  --- Last 5 rows ---")
        for row_idx, row in enumerate(ws.iter_rows(min_row=max(1, ws.max_row-4), max_row=ws.max_row, values_only=False), max(1, ws.max_row-4)):
            vals = []
            for cell in row:
                v = cell.value
                if v is not None:
                    s = str(v)
                    if len(s) > 50:
                        s = s[:50] + "..."
                    vals.append(f"{cell.column_letter}{cell.row}={s}")
            if vals:
                print(f"  R{row_idx:3d}: {' | '.join(vals[:15])}")

wb.close()
print("\n=== ANALYSIS COMPLETE ===")
