#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
대시보드 스킬 모듈

스킬:
  - run_issue_trend: PSRC 이슈 트렌드 대시보드 생성 (Excel)
"""

from __future__ import annotations

import os
import re
import traceback
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

import yaml

from scripts.telegram.skill_utils import (
    ISSUES_DIR,
    load_vault_issues,
)


def _parse_frontmatter(filepath: Path) -> dict:
    """마크다운 파일에서 YAML frontmatter 파싱."""
    try:
        text = filepath.read_text(encoding="utf-8")
        if not text.startswith("---"):
            return {}
        parts = text.split("---", 2)
        if len(parts) < 3:
            return {}
        return yaml.safe_load(parts[1]) or {}
    except Exception:
        return {}


def _parse_date(val) -> Optional[datetime]:
    """다양한 날짜 형식 파싱."""
    if isinstance(val, datetime):
        return val
    if hasattr(val, "isoformat"):  # date object
        return datetime(val.year, val.month, val.day)
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    return None


def _iso_week(dt: datetime) -> str:
    """datetime → 'W07' 형태의 ISO 주차 문자열."""
    return f"W{dt.isocalendar()[1]:02d}"


def run_issue_trend(context: dict) -> dict:
    """PSRC 이슈 트렌드 대시보드 생성.

    Obsidian 01-Issues/*.md 전체를 파싱하여:
    - 카테고리별 분포
    - 우선순위별 분포
    - 주간 추이 (신규 발생)
    - 상태별 분포
    를 Excel 파일로 생성한다.
    """
    send_progress = context.get("send_progress", lambda x: None)
    task_dir = context.get("task_dir", "")

    send_progress("📊 Obsidian 이슈 파일 스캔 중...")

    if not ISSUES_DIR.exists():
        return {
            "result_text": "⚠️ 이슈 디렉토리를 찾을 수 없습니다.",
            "files": [],
        }

    # 모든 이슈 파일에서 frontmatter 수집
    issues = []
    md_files = list(ISSUES_DIR.glob("*.md"))
    for fp in md_files:
        fm = _parse_frontmatter(fp)
        if fm:
            fm["_filename"] = fp.name
            issues.append(fm)

    if not issues:
        return {"result_text": "⚠️ 파싱 가능한 이슈가 없습니다.", "files": []}

    send_progress(f"📊 이슈 {len(issues)}건 분석 중...")

    # 분석 데이터 수집
    category_counter: Counter = Counter()
    priority_counter: Counter = Counter()
    status_counter: Counter = Counter()
    weekly_counter: Counter = Counter()
    owner_counter: Counter = Counter()

    now = datetime.now()

    for issue in issues:
        cat = (issue.get("category") or "미분류").strip().lower()
        pri = (issue.get("priority") or "normal").strip().lower()
        status = (issue.get("status") or "open").strip().lower()
        owner = (issue.get("owner") or "Unassigned").strip()

        category_counter[cat] += 1
        priority_counter[pri] += 1
        status_counter[status] += 1
        owner_counter[owner] += 1

        # 생성일 기반 주간 추이
        created = _parse_date(issue.get("created") or issue.get("date"))
        if created:
            weekly_counter[_iso_week(created)] += 1

    # Excel 생성
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Sheet 1: 요약
        ws_summary = wb.active
        ws_summary.title = "요약"
        ws_summary.append(["P5 이슈 트렌드 대시보드", "", f"생성: {now.strftime('%Y-%m-%d %H:%M')}"])
        ws_summary.append([])
        ws_summary.append(["전체 이슈", len(issues)])
        ws_summary.append(["카테고리 수", len(category_counter)])
        ws_summary.append([])
        ws_summary.append(["카테고리", "건수", "비율"])
        for cat, cnt in category_counter.most_common():
            pct = f"{cnt/len(issues)*100:.1f}%"
            ws_summary.append([cat, cnt, pct])

        # Sheet 2: 카테고리별 분포
        ws_cat = wb.create_sheet("카테고리별")
        ws_cat.append(["카테고리", "건수"])
        for cat, cnt in category_counter.most_common():
            ws_cat.append([cat, cnt])

        # 차트 추가
        if len(category_counter) > 0:
            chart = BarChart()
            chart.title = "카테고리별 이슈 분포"
            chart.y_axis.title = "건수"
            data_ref = Reference(ws_cat, min_col=2, min_row=1, max_row=len(category_counter) + 1)
            cats_ref = Reference(ws_cat, min_col=1, min_row=2, max_row=len(category_counter) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            ws_cat.add_chart(chart, "D2")

        # Sheet 3: 우선순위별
        ws_pri = wb.create_sheet("우선순위별")
        ws_pri.append(["우선순위", "건수"])
        for pri, cnt in priority_counter.most_common():
            ws_pri.append([pri, cnt])

        # Sheet 4: 주간 추이
        ws_weekly = wb.create_sheet("주간추이")
        ws_weekly.append(["주차", "신규 건수"])
        for week in sorted(weekly_counter.keys()):
            ws_weekly.append([week, weekly_counter[week]])

        if len(weekly_counter) > 1:
            chart2 = BarChart()
            chart2.title = "주간 이슈 발생 추이"
            chart2.y_axis.title = "건수"
            data_ref2 = Reference(ws_weekly, min_col=2, min_row=1, max_row=len(weekly_counter) + 1)
            cats_ref2 = Reference(ws_weekly, min_col=1, min_row=2, max_row=len(weekly_counter) + 1)
            chart2.add_data(data_ref2, titles_from_data=True)
            chart2.set_categories(cats_ref2)
            ws_weekly.add_chart(chart2, "D2")

        # Sheet 5: 담당자별
        ws_owner = wb.create_sheet("담당자별")
        ws_owner.append(["담당자", "건수"])
        for owner, cnt in owner_counter.most_common(15):
            ws_owner.append([owner, cnt])

        # 저장
        out_path = os.path.join(task_dir, f"P5_이슈트렌드_{now.strftime('%Y%m%d')}.xlsx")
        wb.save(out_path)

        # 텍스트 요약 구성
        top3_cats = category_counter.most_common(3)
        top3_text = ", ".join(f"{c}({n}건)" for c, n in top3_cats)

        summary = (
            f"📊 **PSRC 이슈 트렌드 대시보드**\n\n"
            f"총 이슈: {len(issues)}건\n"
            f"카테고리: {len(category_counter)}종\n"
            f"Top 3: {top3_text}\n\n"
            f"**우선순위 분포:**\n"
        )
        for pri, cnt in priority_counter.most_common():
            summary += f"  {pri}: {cnt}건\n"

        summary += f"\n**상태 분포:**\n"
        for st, cnt in status_counter.most_common():
            summary += f"  {st}: {cnt}건\n"

        return {"result_text": summary, "files": [out_path]}

    except ImportError:
        # openpyxl 없으면 텍스트만 반환
        top3_cats = category_counter.most_common(3)
        top3_text = ", ".join(f"{c}({n}건)" for c, n in top3_cats)

        summary = (
            f"📊 **PSRC 이슈 트렌드 분석** (텍스트)\n\n"
            f"총 이슈: {len(issues)}건\n"
            f"Top 3 카테고리: {top3_text}\n\n"
            f"**카테고리별:**\n"
        )
        for cat, cnt in category_counter.most_common():
            bar = "█" * min(cnt // 5, 20)
            summary += f"  {cat}: {cnt}건 {bar}\n"

        summary += f"\n**우선순위별:**\n"
        for pri, cnt in priority_counter.most_common():
            summary += f"  {pri}: {cnt}건\n"

        summary += (
            f"\n⚠️ Excel 생성에는 openpyxl이 필요합니다: "
            f"`pip install openpyxl`"
        )
        return {"result_text": summary, "files": []}

    except Exception as e:
        return {
            "result_text": f"❌ 트렌드 분석 오류: {e}\n{traceback.format_exc()[-500:]}",
            "files": [],
        }
