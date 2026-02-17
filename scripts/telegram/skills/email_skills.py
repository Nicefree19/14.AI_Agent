#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
이메일 강화 스킬 모듈

- email_attachment: 이메일 첨부파일 다운로드 + 자동 분석
- email_send: Outlook으로 새 이메일 작성/발송 (2단계 확인)
- email_reply: 수신 메일에 답장/전체답장 (2단계 확인)
"""

from __future__ import annotations

import json
import logging
import os
import re
import sys
import time
import traceback
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# scripts/ 디렉토리 import path
_SCRIPTS_DIR = str(Path(__file__).resolve().parent.parent.parent)
if _SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, _SCRIPTS_DIR)

log = logging.getLogger(__name__)

# ── 드래프트 저장소 ─────────────────────────────────────────────
_DRAFT_DIR = str(
    Path(__file__).resolve().parent.parent.parent.parent / "telegram_data" / "email_drafts"
)
_DRAFT_EXPIRY_MINUTES = 10

# ── 키워드 패턴 ─────────────────────────────────────────────────
_ATTACH_KEYWORDS = [
    "첨부파일분석", "첨부파일확인", "메일첨부분석", "메일첨부파일",
    "이메일첨부", "첨부확인", "첨부분석", "첨부다운",
    "첨부파일", "첨부", "attachment",
]
_SEND_KEYWORDS = [
    "메일발송", "메일보내", "메일전송", "메일작성",
    "이메일보내", "이메일발송", "이메일전송",
    "발송", "보내", "전송", "작성",
]
_REPLY_KEYWORDS = [
    "메일회신", "메일답장", "전체회신", "메일답변",
    "회신", "답장", "답신", "reply",
]

_CONFIRM_KEYWORDS = [
    "보내", "발송", "전송", "확인", "ok", "yes", "네", "ㅇㅇ",
    "보내줘", "발송해", "전송해", "보내세요", "발송하세요",
    "send", "confirm",
]


def _strip_keywords(text: str, keywords: List[str]) -> str:
    """텍스트에서 트리거 키워드를 제거하여 실제 내용 추출."""
    result = text.lower().strip()
    for kw in sorted(keywords, key=len, reverse=True):
        result = result.replace(kw, "")
    result = re.sub(r'\s+', ' ', result).strip()
    return result


# ═══════════════════════════════════════════════════════════════
#  헬퍼: 드래프트 관리
# ═══════════════════════════════════════════════════════════════

def _ensure_draft_dir() -> str:
    """드래프트 디렉토리 생성 및 반환."""
    os.makedirs(_DRAFT_DIR, exist_ok=True)
    return _DRAFT_DIR


def _save_draft(draft_data: dict) -> str:
    """드래프트 JSON 저장. 반환: 파일 경로."""
    d = _ensure_draft_dir()
    ts = int(time.time())
    path = os.path.join(d, f"draft_{ts}.json")
    draft_data["created_at"] = datetime.now().isoformat()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(draft_data, f, ensure_ascii=False, indent=2)
    return path


def _load_latest_draft(draft_type: str) -> Optional[dict]:
    """최신 유효 드래프트 로드. 만료된 것은 삭제."""
    d = _ensure_draft_dir()
    drafts = sorted(Path(d).glob("draft_*.json"), reverse=True)
    cutoff = datetime.now() - timedelta(minutes=_DRAFT_EXPIRY_MINUTES)

    for dp in drafts:
        try:
            with open(dp, "r", encoding="utf-8") as f:
                data = json.load(f)
            created = datetime.fromisoformat(data.get("created_at", "2000-01-01"))
            if created < cutoff:
                dp.unlink(missing_ok=True)
                continue
            if data.get("type") == draft_type:
                dp.unlink(missing_ok=True)  # 사용 후 삭제
                return data
        except Exception:
            continue
    return None


def _clear_old_drafts():
    """만료된 드래프트 정리."""
    d = _ensure_draft_dir()
    cutoff = datetime.now() - timedelta(minutes=_DRAFT_EXPIRY_MINUTES)
    for dp in Path(d).glob("draft_*.json"):
        try:
            with open(dp, "r", encoding="utf-8") as f:
                data = json.load(f)
            created = datetime.fromisoformat(data.get("created_at", "2000-01-01"))
            if created < cutoff:
                dp.unlink(missing_ok=True)
        except Exception:
            dp.unlink(missing_ok=True)


# ═══════════════════════════════════════════════════════════════
#  헬퍼: 발송 확인 감지
# ═══════════════════════════════════════════════════════════════

def _detect_confirmation(instruction: str) -> bool:
    """사용자 입력이 발송 확인인지 감지."""
    text = instruction.strip().lower()
    # 짧은 확인 메시지 (10자 이하)
    if len(text) <= 10:
        for kw in _CONFIRM_KEYWORDS:
            if kw in text:
                return True
    return False


# ═══════════════════════════════════════════════════════════════
#  헬퍼: 대상 이메일 검색
# ═══════════════════════════════════════════════════════════════

def _find_target_email(
    instruction: str,
    send_progress,
    require_attachments: bool = False,
    limit: int = 5,
) -> Tuple[Optional[list], Optional[str]]:
    """지시문에서 대상 이메일을 검색하여 반환.

    Returns:
        (messages_list, error_text)
    """
    try:
        from adapters.outlook_adapter import OutlookAdapter
    except ImportError:
        return None, "⚠️ Outlook 어댑터를 불러올 수 없습니다."

    adapter = OutlookAdapter()
    if not adapter.initialize():
        return None, "⚠️ Outlook에 연결할 수 없습니다."

    # 검색 파라미터 추출
    text = instruction.lower()
    sender = None
    subject_kw = None
    search_kw = None

    # "마지막 메일" / "최근 메일"
    if "마지막" in text or "최근" in text or "last" in text:
        msgs = adapter.search_emails(
            limit=1,
            days_back=30,
            has_attachments=require_attachments,
        )
        if msgs:
            return msgs, None
        return None, "⚠️ 최근 30일 내 메일을 찾을 수 없습니다."

    # 발신자 추출: "김과장" "from:xxx" 패턴
    from_match = re.search(r'(?:from:|발신:|보낸사람:)\s*(\S+)', text)
    if from_match:
        sender = from_match.group(1)

    # 제목 추출: "제목:xxx" 패턴
    subj_match = re.search(r'(?:subject:|제목:)\s*(.+?)(?:\s*[-—]|$)', text)
    if subj_match:
        subject_kw = subj_match.group(1).strip()

    # 나머지 텍스트를 검색 키워드로
    cleaned = _strip_keywords(instruction, _ATTACH_KEYWORDS + _REPLY_KEYWORDS + _SEND_KEYWORDS)
    cleaned = re.sub(r'(?:from:|발신:|보낸사람:|subject:|제목:)\s*\S+', '', cleaned).strip()
    if cleaned and len(cleaned) >= 2:
        search_kw = cleaned

    msgs = adapter.search_emails(
        sender=sender,
        subject=subject_kw,
        keyword=search_kw,
        limit=limit,
        days_back=14,
        has_attachments=require_attachments,
    )

    if not msgs:
        return None, "⚠️ 조건에 맞는 이메일을 찾을 수 없습니다."

    return msgs, None


# ═══════════════════════════════════════════════════════════════
#  헬퍼: 수신자 이름 → 이메일 주소 변환
# ═══════════════════════════════════════════════════════════════

def _resolve_recipient(name: str) -> Optional[str]:
    """최근 메일에서 이름으로 이메일 주소 검색."""
    try:
        from adapters.outlook_adapter import OutlookAdapter
        adapter = OutlookAdapter()
        if not adapter.initialize():
            return None

        # 받은편지함 + 보낸편지함에서 검색
        for folder in ["inbox", "sent"]:
            msgs = adapter.search_emails(
                sender=name if folder == "inbox" else None,
                keyword=name if folder == "sent" else None,
                folder=folder,
                limit=5,
                days_back=60,
            )
            for msg in msgs:
                # 이메일 주소 추출
                email_match = re.search(r'<([^>]+@[^>]+)>', msg.sender)
                if email_match:
                    return email_match.group(1)
                # sender 자체가 이메일인 경우
                if "@" in msg.sender:
                    return msg.sender.strip()
        return None
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════
#  헬퍼: 이메일 작성 파싱
# ═══════════════════════════════════════════════════════════════

def _parse_email_compose(instruction: str) -> Dict[str, str]:
    """지시문에서 수신자/제목/본문 파싱.

    Returns:
        {"to": str, "cc": str, "subject": str, "body": str}
    """
    result = {"to": "", "cc": "", "subject": "", "body": ""}
    text = instruction.strip()

    # 수신자: "to:xxx" 또는 "수신:xxx" 또는 "xxx한테" 또는 "xxx에게"
    to_match = re.search(
        r'(?:to:|수신:|받는사람:)\s*([^\s,;]+(?:@[^\s,;]+)?)',
        text, re.IGNORECASE,
    )
    if to_match:
        result["to"] = to_match.group(1)
    else:
        # "김과장한테" / "XXX에게" 패턴
        name_match = re.search(r'(\S{2,6})(?:한테|에게|님한테|님에게)', text)
        if name_match:
            name = name_match.group(1)
            resolved = _resolve_recipient(name)
            if resolved:
                result["to"] = resolved
            else:
                result["to"] = name  # 주소록에서 못 찾으면 이름 그대로

    # CC: "cc:xxx"
    cc_match = re.search(r'(?:cc:|참조:)\s*([^\s,;]+)', text, re.IGNORECASE)
    if cc_match:
        result["cc"] = cc_match.group(1)

    # 제목: "제목:xxx" 또는 "subject:xxx"
    subj_match = re.search(
        r'(?:subject:|제목:)\s*(.+?)(?:\s*[-—]\s*|$)',
        text, re.IGNORECASE,
    )
    if subj_match:
        result["subject"] = subj_match.group(1).strip()

    # 본문: "내용:" 이후 또는 "-" 이후 텍스트
    body_match = re.search(
        r'(?:body:|본문:|내용:)\s*(.+)',
        text, re.IGNORECASE | re.DOTALL,
    )
    if body_match:
        result["body"] = body_match.group(1).strip()
    elif not result["body"]:
        # 키워드 제거 후 남은 텍스트를 본문으로
        cleaned = _strip_keywords(text, _SEND_KEYWORDS)
        # 수신자/제목 부분 제거
        cleaned = re.sub(
            r'(?:to:|수신:|받는사람:|cc:|참조:|subject:|제목:)\s*\S+',
            '', cleaned,
        )
        cleaned = re.sub(r'\S{2,6}(?:한테|에게|님한테|님에게)', '', cleaned)
        cleaned = cleaned.strip()
        if cleaned:
            # 제목이 없으면 첫 줄을 제목으로
            if not result["subject"]:
                lines = cleaned.split("\n", 1)
                result["subject"] = lines[0].strip()[:50]
                if len(lines) > 1:
                    result["body"] = lines[1].strip()
            else:
                result["body"] = cleaned

    return result


# ═══════════════════════════════════════════════════════════════
#  헬퍼: 회신 내용 파싱
# ═══════════════════════════════════════════════════════════════

def _parse_reply_content(instruction: str) -> Tuple[str, str, bool]:
    """지시문에서 검색 쿼리, 회신 본문, reply_all 추출.

    Returns:
        (search_query, reply_body, reply_all)
    """
    text = instruction.strip()
    reply_all = False

    # reply_all 감지
    if "전체회신" in text or "reply all" in text.lower() or "전체답장" in text:
        reply_all = True

    # "- " 또는 ":" 구분자로 검색어/본문 분리
    # 예: "SEN-070 메일 회신 - 검토중입니다"
    # 예: "마지막 메일에 답장: 확인했습니다"
    parts = re.split(r'\s*[-—:]\s*', text, maxsplit=1)

    if len(parts) == 2:
        search_part = _strip_keywords(parts[0], _REPLY_KEYWORDS)
        body_part = parts[1].strip()
    else:
        # 분리 불가 → 전체를 검색어로, 본문 비움
        search_part = _strip_keywords(text, _REPLY_KEYWORDS)
        body_part = ""

    # 검색어 정리
    search_part = re.sub(r'(?:메일|이메일|에|에게|한테)\s*(?:답장|회신|답신)', '', search_part)
    search_part = search_part.strip()

    return search_part, body_part, reply_all


# ═══════════════════════════════════════════════════════════════
#  1. 첨부파일 분석
# ═══════════════════════════════════════════════════════════════

def run_email_attachment(context: dict) -> dict:
    """이메일 첨부파일 다운로드 + 자동 분석.

    Flow:
    1. 지시문에서 대상 이메일 파악
    2. search_emails → has_attachments 필터
    3. get_attachments → 파일 다운로드
    4. 확장자별 자동 분석 (PDF/Excel 재사용)
    5. 통합 분석 결과 반환
    """
    send_progress = context.get("send_progress", lambda x: None)
    task_dir = context.get("task_dir", ".")
    combined = context.get("combined", {})
    instruction = combined.get("combined_instruction", "")

    send_progress("📎 이메일 첨부파일 분석 중...")

    # 1. 대상 이메일 검색
    msgs, error = _find_target_email(
        instruction, send_progress, require_attachments=True, limit=5,
    )

    if error:
        return {"result_text": error, "files": []}

    if not msgs:
        return {
            "result_text": "⚠️ 첨부파일이 있는 이메일을 찾을 수 없습니다.",
            "files": [],
        }

    # 복수 매칭 → 목록 표시 (3개 이상이면 선택 유도)
    if len(msgs) > 1:
        lines = [
            "📎 첨부파일이 있는 이메일 목록",
            "━" * 30,
        ]
        for i, msg in enumerate(msgs[:5], 1):
            att_names = msg.raw_metadata.get("attachment_names", [])
            ts = msg.timestamp.strftime("%m/%d %H:%M") if msg.timestamp else ""
            lines.append(
                f"{i}. [{ts}] {msg.sender.split('<')[0].strip()}"
            )
            lines.append(f"   제목: {msg.subject[:40]}")
            lines.append(f"   첨부: {', '.join(att_names[:3])}")
            lines.append("")

        lines.append('💡 "첫번째 메일 첨부 분석" 또는 "마지막 메일 첨부 분석"으로 선택하세요.')

        # 1개만 있으면 바로 분석
        if len(msgs) == 1:
            pass  # 아래로 fall through
        else:
            return {"result_text": "\n".join(lines), "files": []}

    # 단일 메일 → 첨부파일 다운로드 + 분석
    target_msg = msgs[0]
    send_progress(f"📥 첨부파일 다운로드 중... ({target_msg.subject[:30]})")

    try:
        from adapters.outlook_adapter import OutlookAdapter
        adapter = OutlookAdapter()
        if not adapter.initialize():
            return {"result_text": "⚠️ Outlook 연결 실패", "files": []}

        output_dir = Path(task_dir) / "attachments"
        saved_files = adapter.get_attachments(target_msg.id, output_dir)

        if not saved_files:
            return {
                "result_text": "⚠️ 첨부파일 다운로드에 실패했습니다.",
                "files": [],
            }

    except Exception as e:
        return {
            "result_text": f"❌ 첨부파일 다운로드 오류: {e}",
            "files": [],
        }

    # 4. 확장자별 자동 분석
    send_progress(f"🔍 {len(saved_files)}개 첨부파일 분석 중...")

    analysis_lines = [
        f"📎 이메일 첨부파일 분석 결과",
        f"━" * 30,
        f"📧 발신: {target_msg.sender.split('<')[0].strip()}",
        f"📋 제목: {target_msg.subject}",
        f"📅 날짜: {target_msg.timestamp.strftime('%Y-%m-%d %H:%M') if target_msg.timestamp else 'N/A'}",
        f"📎 첨부: {len(saved_files)}개",
        "",
    ]

    output_files = [str(f) for f in saved_files]

    for fpath in saved_files:
        ext = fpath.suffix.lower()
        fname = fpath.name
        fsize = fpath.stat().st_size if fpath.exists() else 0

        analysis_lines.append(f"── {fname} ({_format_size(fsize)}) ──")

        if ext == ".pdf":
            # PDF 분석 재사용
            try:
                sub_result = _analyze_pdf_file(str(fpath), send_progress)
                analysis_lines.append(sub_result)
            except Exception as e:
                analysis_lines.append(f"  PDF 분석 오류: {e}")

        elif ext in (".xlsx", ".xls", ".csv"):
            # 엑셀 분석 재사용
            try:
                sub_result = _analyze_excel_file(str(fpath), send_progress)
                analysis_lines.append(sub_result)
            except Exception as e:
                analysis_lines.append(f"  엑셀 분석 오류: {e}")

        elif ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp"):
            analysis_lines.append(f"  이미지 파일 ({ext})")

        elif ext in (".docx", ".doc"):
            analysis_lines.append(f"  Word 문서 ({ext})")

        else:
            analysis_lines.append(f"  파일 형식: {ext}")

        analysis_lines.append("")

    return {
        "result_text": "\n".join(analysis_lines),
        "files": output_files,
    }


def _format_size(size_bytes: int) -> str:
    """바이트를 읽기 좋은 크기로 변환."""
    if size_bytes < 1024:
        return f"{size_bytes}B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f}KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f}MB"


def _analyze_pdf_file(fpath: str, send_progress) -> str:
    """PDF 파일 간이 분석. skill_utils의 extract_pdf_text 재사용."""
    try:
        from telegram.skill_utils import extract_pdf_text
        text, pages = extract_pdf_text(fpath, max_pages=10)
        if not text or pages == 0:
            return "  텍스트 추출 실패"

        preview = text[:500].replace("\n", " ")
        return (
            f"  페이지: {pages}p | 텍스트: {len(text):,}자\n"
            f"  미리보기: {preview}..."
        )
    except Exception as e:
        return f"  PDF 분석 실패: {e}"


def _analyze_excel_file(fpath: str, send_progress) -> str:
    """엑셀 파일 간이 분석."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        sheets_info = []
        for name in wb.sheetnames[:5]:
            ws = wb[name]
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            sheets_info.append(f"  {name}: {rows}행 x {cols}열")
        wb.close()
        return "\n".join(sheets_info) if sheets_info else "  시트 정보 없음"
    except Exception:
        # openpyxl 실패 시 pandas 시도
        try:
            import pandas as pd
            if fpath.endswith(".csv"):
                df = pd.read_csv(fpath, nrows=5)
            else:
                df = pd.read_excel(fpath, nrows=5)
            return f"  {len(df)}행 x {len(df.columns)}열 | 컬럼: {', '.join(df.columns[:5])}"
        except Exception as e2:
            return f"  엑셀 분석 실패: {e2}"


# ═══════════════════════════════════════════════════════════════
#  2. 이메일 발송 (2단계 확인)
# ═══════════════════════════════════════════════════════════════

def run_email_send(context: dict) -> dict:
    """이메일 발송 — 2단계 확인 (미리보기 → 확인 후 발송).

    Flow:
    1. 기존 draft 확인 → 확인 키워드 감지 → 즉시 발송
    2. draft 없으면 → 수신자/제목/본문 파싱 → 미리보기 → draft 저장
    """
    send_progress = context.get("send_progress", lambda x: None)
    task_dir = context.get("task_dir", ".")
    combined = context.get("combined", {})
    instruction = combined.get("combined_instruction", "")

    # 만료 드래프트 정리
    _clear_old_drafts()

    # 1. 기존 draft 확인 + 확인 키워드 감지
    if _detect_confirmation(instruction):
        draft = _load_latest_draft("send")
        if draft:
            send_progress("📤 이메일 발송 중...")
            return _execute_send(draft, send_progress)

    # 2. 새 이메일 작성 → 미리보기
    send_progress("📝 이메일 작성 중...")
    parsed = _parse_email_compose(instruction)

    if not parsed["to"]:
        return {
            "result_text": (
                "⚠️ 수신자를 지정해주세요.\n\n"
                "사용법:\n"
                '• "메일보내 to:email@example.com 제목:회의록 내용:첨부합니다"\n'
                '• "김과장한테 이슈현황 메일 보내줘"\n'
                '• "수신:team@company.com 주간보고 전달합니다"'
            ),
            "files": [],
        }

    # task_dir 내 파일을 첨부 후보로 검색
    attach_candidates = []
    if os.path.isdir(task_dir):
        for f in os.listdir(task_dir):
            fpath = os.path.join(task_dir, f)
            if os.path.isfile(fpath) and not f.startswith("."):
                attach_candidates.append(fpath)

    # 미리보기 구성
    preview_lines = [
        "📧 이메일 발송 미리보기",
        "━" * 30,
        f"📤 수신: {parsed['to']}",
    ]
    if parsed["cc"]:
        preview_lines.append(f"📋 참조: {parsed['cc']}")
    preview_lines.append(f"📋 제목: {parsed['subject'] or '(제목 없음)'}")
    preview_lines.append(f"📝 내용: {parsed['body'][:200] or '(내용 없음)'}")
    if attach_candidates:
        fnames = [os.path.basename(f) for f in attach_candidates[:5]]
        preview_lines.append(f"📎 첨부 후보: {', '.join(fnames)}")
    preview_lines.append("")
    preview_lines.append('⚠️ "보내줘" 또는 "발송"으로 발송을 확인해주세요.')

    # draft 저장
    draft_data = {
        "type": "send",
        "to": parsed["to"],
        "cc": parsed["cc"],
        "subject": parsed["subject"],
        "body": parsed["body"],
        "attachments": attach_candidates[:5] if attach_candidates else [],
    }
    _save_draft(draft_data)

    return {
        "result_text": "\n".join(preview_lines),
        "files": [],
    }


def _execute_send(draft: dict, send_progress) -> dict:
    """실제 이메일 발송 실행."""
    try:
        from adapters.outlook_adapter import OutlookAdapter
        adapter = OutlookAdapter()
        if not adapter.initialize():
            return {"result_text": "⚠️ Outlook 연결 실패", "files": []}

        result = adapter.send_email(
            to=draft["to"],
            subject=draft.get("subject", ""),
            body=draft.get("body", ""),
            cc=draft.get("cc") or None,
            attachments=draft.get("attachments") or None,
        )

        if result.get("success"):
            att_count = result.get("attachment_count", 0)
            att_info = f" (첨부 {att_count}개)" if att_count else ""
            return {
                "result_text": (
                    f"✅ 이메일 발송 완료!\n"
                    f"📤 수신: {result['to']}\n"
                    f"📋 제목: {result['subject']}{att_info}"
                ),
                "files": [],
            }
        else:
            return {
                "result_text": f"❌ 이메일 발송 실패: {result.get('error', '알 수 없는 오류')}",
                "files": [],
            }

    except Exception as e:
        return {
            "result_text": f"❌ 이메일 발송 오류: {e}\n{traceback.format_exc()[-500:]}",
            "files": [],
        }


# ═══════════════════════════════════════════════════════════════
#  3. 이메일 회신 (2단계 확인)
# ═══════════════════════════════════════════════════════════════

def run_email_reply(context: dict) -> dict:
    """이메일 회신 — 2단계 확인 (미리보기 → 확인 후 발송).

    Flow:
    1. 기존 draft 확인 → 확인 키워드 감지 → 즉시 회신
    2. draft 없으면 → 대상 이메일 검색 + 회신 내용 파싱 → 미리보기
    """
    send_progress = context.get("send_progress", lambda x: None)
    task_dir = context.get("task_dir", ".")
    combined = context.get("combined", {})
    instruction = combined.get("combined_instruction", "")

    # 만료 드래프트 정리
    _clear_old_drafts()

    # 1. 기존 draft 확인 + 확인 키워드
    if _detect_confirmation(instruction):
        draft = _load_latest_draft("reply")
        if draft:
            send_progress("📤 이메일 회신 중...")
            return _execute_reply(draft, send_progress)

    # 2. 새 회신 준비
    send_progress("📧 회신 대상 이메일 검색 중...")

    search_query, reply_body, reply_all = _parse_reply_content(instruction)

    # 대상 이메일 검색
    msgs, error = _find_target_email(
        search_query or instruction, send_progress, require_attachments=False, limit=3,
    )

    if error:
        return {"result_text": error, "files": []}

    if not msgs:
        return {
            "result_text": (
                "⚠️ 회신할 이메일을 찾을 수 없습니다.\n\n"
                "사용법:\n"
                '• "마지막 메일에 답장 - 확인했습니다"\n'
                '• "SEN-070 메일 회신 - 검토중입니다"\n'
                '• "김과장 메일 전체회신 - 일정 조정 부탁합니다"'
            ),
            "files": [],
        }

    # 복수 매칭 → 첫 번째 사용 (가장 최근)
    target_msg = msgs[0]

    if not reply_body:
        # 본문이 비어있으면 안내
        return {
            "result_text": (
                f"📧 회신 대상 이메일:\n"
                f"  발신: {target_msg.sender.split('<')[0].strip()}\n"
                f"  제목: {target_msg.subject}\n"
                f"  날짜: {target_msg.timestamp.strftime('%m/%d %H:%M') if target_msg.timestamp else 'N/A'}\n\n"
                f'⚠️ 회신 내용을 지정해주세요.\n'
                f'예: "이 메일에 답장 - 확인했습니다. 다음 주 회의에서 논의하겠습니다."'
            ),
            "files": [],
        }

    # 미리보기 구성
    reply_type = "전체회신" if reply_all else "회신"
    preview_lines = [
        f"📧 이메일 {reply_type} 미리보기",
        "━" * 30,
        f"📥 원본 발신: {target_msg.sender.split('<')[0].strip()}",
        f"📋 원본 제목: {target_msg.subject}",
        f"📤 {reply_type} 대상: {'전체' if reply_all else target_msg.sender.split('<')[0].strip()}",
        f"📝 회신 내용: {reply_body[:200]}",
        "",
        f'⚠️ "보내줘" 또는 "발송"으로 회신을 확인해주세요.',
    ]

    # draft 저장
    draft_data = {
        "type": "reply",
        "message_id": target_msg.id,
        "reply_all": reply_all,
        "body": reply_body,
        "original_sender": target_msg.sender,
        "original_subject": target_msg.subject,
        "attachments": [],
    }
    _save_draft(draft_data)

    return {
        "result_text": "\n".join(preview_lines),
        "files": [],
    }


def _execute_reply(draft: dict, send_progress) -> dict:
    """실제 이메일 회신 실행."""
    try:
        from adapters.outlook_adapter import OutlookAdapter
        adapter = OutlookAdapter()
        if not adapter.initialize():
            return {"result_text": "⚠️ Outlook 연결 실패", "files": []}

        result = adapter.reply_email(
            message_id=draft["message_id"],
            body=draft["body"],
            reply_all=draft.get("reply_all", False),
            attachments=draft.get("attachments") or None,
        )

        if result.get("success"):
            reply_type = "전체회신" if result.get("reply_all") else "회신"
            return {
                "result_text": (
                    f"✅ 이메일 {reply_type} 완료!\n"
                    f"📤 수신: {result['to']}\n"
                    f"📋 제목: {result['subject']}"
                ),
                "files": [],
            }
        else:
            return {
                "result_text": f"❌ 이메일 회신 실패: {result.get('error', '알 수 없는 오류')}",
                "files": [],
            }

    except Exception as e:
        return {
            "result_text": f"❌ 이메일 회신 오류: {e}\n{traceback.format_exc()[-500:]}",
            "files": [],
        }
