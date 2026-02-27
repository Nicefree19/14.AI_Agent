"""
Drawing Report Generator — 도면 분석 결과 리포트 생성

분석 결과를 텔레그램 메시지용 구조화된 텍스트로 변환.
Excel 출력 옵션 지원.
"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

log = logging.getLogger(__name__)


class DrawingReportGenerator:
    """도면 분석 결과 리포트 생성."""

    # 텔레그램 메시지 최대 길이 (4096자, 여유분 포함)
    MAX_MESSAGE_LEN = 3800

    def generate(
        self,
        analysis: Dict[str, Any],
        format: str = "text",
    ) -> str:
        """분석 결과를 텔레그램 메시지용 텍스트로 변환.

        Args:
            analysis: StructuralAnalyzer.analyze() 반환값
            format: "text" (기본) 또는 "brief" (요약)

        Returns:
            구조화된 텍스트 리포트
        """
        lines: List[str] = []

        file_name = analysis.get("file_name", "도면")
        pages = analysis.get("total_pages", 1)

        lines.append(f"📐 도면 정밀 분석: {file_name}")
        lines.append("━━━━━━━━━━━━━━━━━━━━━━━━")

        if pages > 1:
            lines.append(f"📄 총 페이지: {pages}")

        # ── 1. 기본 정보 (타이틀 블록) ──
        title_block = analysis.get("title_block", {})
        if title_block:
            lines.append(self._section_title_block(title_block))

        # ── 2. 치수 정보 ──
        dimensions = analysis.get("dimensions", [])
        if dimensions:
            lines.append(self._section_dimensions(dimensions, format))

        # ── 3. 구조 상세 ──
        members = analysis.get("members", [])
        if members:
            lines.append(self._section_structural(members, format))

        # ── 4. 그리드 시스템 ──
        grid = analysis.get("grid", {})
        if grid:
            lines.append(self._section_grid(grid))

        # ── 5. 주석/사양 ──
        notes = analysis.get("notes", [])
        specs = analysis.get("specifications", [])
        standards = analysis.get("standards", [])
        if notes or specs or standards:
            lines.append(self._section_notes(notes, specs, standards, format))

        # ── 6. 수량 집계 ──
        quantities = analysis.get("quantities", {})
        if quantities:
            lines.append(self._section_quantities(quantities))

        # ── 7. 도면번호 / SEN 이슈 ──
        drawing_refs = analysis.get("drawing_refs", [])
        sen_issues = analysis.get("sen_issues", [])
        if drawing_refs or sen_issues:
            lines.append(self._section_references(drawing_refs, sen_issues))

        # ── 8. 품질 경고 ──
        quality = analysis.get("quality", {})
        if quality:
            lines.append(self._section_quality(quality))

        # ── Zone 분석 요약 ──
        zones_info = analysis.get("zones_detected", [])
        if zones_info:
            lines.append(f"\n🔍 Zone 분석: {len(zones_info)}개 영역 감지")
            for z in zones_info:
                lines.append(f"  • {z.get('type', '?')} ({z.get('confidence', 0):.0%})")

        result = "\n".join(lines)

        # 길이 제한
        if len(result) > self.MAX_MESSAGE_LEN:
            result = result[:self.MAX_MESSAGE_LEN - 50]
            result += "\n\n... (전체 결과는 첨부 파일 참조)"

        return result

    # ─── 섹션별 생성 ──────────────────────────────────────────

    def _section_title_block(self, tb: Dict) -> str:
        """타이틀 블록 섹션."""
        parts = ["\n📋 기본 정보"]

        field_map = {
            "drawing_no": "도면번호",
            "revision": "리비전",
            "date": "날짜",
            "scale": "스케일",
            "designer": "설계자",
            "checker": "검토자",
            "approver": "승인자",
            "project": "프로젝트",
            "title": "도면명",
        }

        for key, label in field_map.items():
            val = tb.get(key, "")
            if val:
                parts.append(f"  • {label}: {val}")

        return "\n".join(parts)

    def _section_dimensions(
        self,
        dims: List[Dict],
        format: str,
    ) -> str:
        """치수 정보 섹션."""
        parts = [f"\n📏 치수 정보 ({len(dims)}개)"]

        # 타입별 그룹화
        by_type: Dict[str, List] = {}
        for d in dims:
            dtype = d.get("type", "기타")
            by_type.setdefault(dtype, []).append(d)

        type_labels = {
            "length": "길이",
            "height": "높이",
            "spacing": "간격",
            "section": "단면",
        }

        limit = 15 if format == "text" else 5

        for dtype, items in by_type.items():
            label = type_labels.get(dtype, dtype)
            parts.append(f"  [{label}]")
            for item in items[:limit]:
                val = item.get("value", "?")
                unit = item.get("unit", "mm")
                name = item.get("label", "")
                suffix = f" ({name})" if name else ""
                parts.append(f"    • {val} {unit}{suffix}")

            if len(items) > limit:
                parts.append(f"    ... 외 {len(items) - limit}개")

        return "\n".join(parts)

    def _section_structural(
        self,
        members: List[Dict],
        format: str,
    ) -> str:
        """구조 상세 섹션."""
        parts = [f"\n🏗️ 구조 상세 ({len(members)}개 부재)"]

        type_labels = {
            "column": "기둥",
            "beam": "보",
            "slab": "슬래브",
            "wall": "벽",
            "footing": "기초",
        }

        limit = 10 if format == "text" else 3

        for member in members[:limit]:
            mtype = type_labels.get(member.get("type", ""), member.get("type", "?"))
            section = member.get("section", "")
            rebar = member.get("rebar", {})
            concrete = member.get("concrete_grade", "")

            line = f"  • {mtype}"
            if section:
                line += f": {section}"
            if concrete:
                line += f" ({concrete})"

            parts.append(line)

            # 철근 상세
            main_rebar = rebar.get("main", "")
            stirrup = rebar.get("stirrup", "")
            if main_rebar:
                parts.append(f"    주근: {main_rebar}")
            if stirrup:
                parts.append(f"    스터럽: {stirrup}")

        if len(members) > limit:
            parts.append(f"  ... 외 {len(members) - limit}개 부재")

        return "\n".join(parts)

    def _section_grid(self, grid: Dict) -> str:
        """그리드 시스템 섹션."""
        parts = ["\n📐 그리드 시스템"]

        grid_x = grid.get("grid_x", [])
        grid_y = grid.get("grid_y", [])

        if grid_x:
            labels = [g.get("label", "?") for g in grid_x]
            parts.append(f"  X축: {', '.join(labels[:15])}")
            # 축간 거리
            distances = [g.get("distance_mm", 0) for g in grid_x if g.get("distance_mm")]
            if distances:
                parts.append(f"  X축 간격: {', '.join(str(d) for d in distances[:10])} mm")

        if grid_y:
            labels = [g.get("label", "?") for g in grid_y]
            parts.append(f"  Y축: {', '.join(labels[:15])}")
            distances = [g.get("distance_mm", 0) for g in grid_y if g.get("distance_mm")]
            if distances:
                parts.append(f"  Y축 간격: {', '.join(str(d) for d in distances[:10])} mm")

        return "\n".join(parts)

    def _section_notes(
        self,
        notes: List[str],
        specs: List[str],
        standards: List[str],
        format: str,
    ) -> str:
        """주석/사양 섹션."""
        parts = ["\n📝 주석 및 사양"]

        limit = 8 if format == "text" else 3

        if notes:
            parts.append(f"  [일반 노트] ({len(notes)}개)")
            for note in notes[:limit]:
                parts.append(f"    • {note[:80]}")

        if specs:
            parts.append(f"  [특기 사양] ({len(specs)}개)")
            for spec in specs[:limit]:
                parts.append(f"    • {spec[:80]}")

        if standards:
            parts.append(f"  [참조 규격]")
            parts.append(f"    {', '.join(standards[:10])}")

        return "\n".join(parts)

    def _section_quantities(self, quantities: Dict) -> str:
        """수량 집계 섹션."""
        parts = ["\n📊 수량 집계"]

        member_counts = quantities.get("member_counts", {})
        if member_counts:
            for mtype, count in member_counts.items():
                parts.append(f"  • {mtype}: {count}개")

        rebar_summary = quantities.get("rebar_summary", "")
        if rebar_summary:
            parts.append(f"  철근 요약: {rebar_summary}")

        total_area = quantities.get("total_area_m2", 0)
        if total_area:
            parts.append(f"  총 면적: {total_area:.1f} m²")

        return "\n".join(parts)

    def _section_references(
        self,
        drawing_refs: List[str],
        sen_issues: List[Dict],
    ) -> str:
        """도면번호 / SEN 이슈 섹션."""
        parts = []

        if drawing_refs:
            parts.append(f"\n🔧 도면번호/부재코드 ({len(drawing_refs)}개)")
            # 카테고리별 분류
            categories: Dict[str, List[str]] = {}
            for ref in drawing_refs:
                prefix = ref.split("-")[0] if "-" in ref else ref[:3]
                categories.setdefault(prefix, []).append(ref)
            for prefix, refs in sorted(categories.items()):
                parts.append(f"  [{prefix}] {', '.join(refs[:10])}")

        if sen_issues:
            parts.append(f"\n📌 관련 SEN 이슈 ({len(sen_issues)}개)")
            for issue in sen_issues[:10]:
                ref = issue.get("ref", "?")
                title = issue.get("title", "")
                priority = issue.get("priority", "?")
                status = issue.get("status", "?")
                if title:
                    parts.append(f"  • {ref}: {title}")
                    parts.append(f"    우선순위: {priority} | 상태: {status}")
                else:
                    parts.append(f"  • {ref} (볼트에 없음)")

        return "\n".join(parts)

    def _section_quality(self, quality: Dict) -> str:
        """품질 경고 섹션."""
        parts = ["\n⚠️ 품질 평가"]

        overall = quality.get("overall_confidence", 0)
        if overall:
            emoji = "🟢" if overall >= 0.8 else "🟡" if overall >= 0.6 else "🔴"
            parts.append(f"  {emoji} 전체 인식 신뢰도: {overall:.0%}")

        dpi = quality.get("estimated_dpi", 0)
        if dpi:
            emoji = "🟢" if dpi >= 200 else "🟡" if dpi >= 150 else "🔴"
            parts.append(f"  {emoji} 추정 DPI: {dpi}")

        warnings = quality.get("warnings", [])
        for w in warnings[:5]:
            parts.append(f"  ⚠️ {w}")

        return "\n".join(parts)

    # ─── Excel 출력 ──────────────────────────────────────────

    def generate_excel(
        self,
        analysis: Dict[str, Any],
        output_path: str,
    ) -> Optional[str]:
        """분석 결과를 Excel 파일로 출력 (선택).

        Returns:
            생성된 Excel 파일 경로 또는 None
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()

            # Sheet 1: 기본 정보
            ws = wb.active
            ws.title = "기본정보"
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font_w = Font(bold=True, size=12, color="FFFFFF")

            # 타이틀
            ws["A1"] = "도면 분석 결과"
            ws["A1"].font = Font(bold=True, size=14)

            tb = analysis.get("title_block", {})
            row = 3
            for key, label in [
                ("drawing_no", "도면번호"), ("revision", "리비전"),
                ("date", "날짜"), ("scale", "스케일"),
                ("designer", "설계자"), ("project", "프로젝트"),
            ]:
                ws.cell(row=row, column=1, value=label).font = header_font
                ws.cell(row=row, column=2, value=tb.get(key, ""))
                row += 1

            # Sheet 2: 치수
            dims = analysis.get("dimensions", [])
            if dims:
                ws2 = wb.create_sheet("치수")
                headers = ["타입", "값", "단위", "라벨"]
                for c, h in enumerate(headers, 1):
                    cell = ws2.cell(row=1, column=c, value=h)
                    cell.font = header_font_w
                    cell.fill = header_fill

                for r, dim in enumerate(dims[:200], 2):
                    ws2.cell(row=r, column=1, value=dim.get("type", ""))
                    ws2.cell(row=r, column=2, value=dim.get("value", ""))
                    ws2.cell(row=r, column=3, value=dim.get("unit", ""))
                    ws2.cell(row=r, column=4, value=dim.get("label", ""))

            # Sheet 3: 구조 부재
            members = analysis.get("members", [])
            if members:
                ws3 = wb.create_sheet("구조부재")
                headers = ["타입", "단면", "주근", "스터럽", "콘크리트"]
                for c, h in enumerate(headers, 1):
                    cell = ws3.cell(row=1, column=c, value=h)
                    cell.font = header_font_w
                    cell.fill = header_fill

                for r, m in enumerate(members[:200], 2):
                    rebar = m.get("rebar", {})
                    ws3.cell(row=r, column=1, value=m.get("type", ""))
                    ws3.cell(row=r, column=2, value=m.get("section", ""))
                    ws3.cell(row=r, column=3, value=rebar.get("main", ""))
                    ws3.cell(row=r, column=4, value=rebar.get("stirrup", ""))
                    ws3.cell(row=r, column=5, value=m.get("concrete_grade", ""))

            # 저장
            wb.save(output_path)
            log.info("Excel 리포트 생성: %s", output_path)
            return output_path

        except Exception as e:
            log.error("Excel 리포트 생성 실패: %s", e)
            return None
